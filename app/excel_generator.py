"""
Excel report generation with 3 sheets: Updated, Current, Past
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict
from pathlib import Path
from datetime import datetime
import re


def extract_resource_id(url: str) -> int:
    """Extract resource ID from URL"""
    match = re.search(r'/exchange/(\d+)/', url)
    if match:
        return int(match.group(1))
    return 999999  # Default for URLs without ID


def sort_by_resource_id(data: List[Dict]) -> List[Dict]:
    """Sort list of resources by resource ID extracted from URL"""
    return sorted(data, key=lambda x: extract_resource_id(x.get('url', '')))


def create_excel_report(updated_results: List[Dict], current_results: List[Dict],
                       past_results: List[Dict], output_path: Path) -> str:
    """
    Create Excel report with 3 sheets.

    Args:
        updated_results: List of new or updated resources
        current_results: All resources from current scrape
        past_results: All resources from previous scrape
        output_path: Path to save the Excel file

    Returns:
        Path to created file
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define headers with Resource ID first
    headers = ['Resource ID', 'Title', 'URL', 'Version', 'Updated Date', 'Developer ID', 'Contributor', 'Tagline']

    # Sort all datasets by Resource ID
    updated_sorted = sort_by_resource_id(updated_results)
    current_sorted = sort_by_resource_id(current_results)
    past_sorted = sort_by_resource_id(past_results)

    # Create Sheet 1: Updated Results
    ws_updated = wb.create_sheet('Updated Results')
    _write_sheet(ws_updated, headers, updated_sorted)

    # Create Sheet 2: Current Results
    ws_current = wb.create_sheet('Current Results')
    _write_sheet(ws_current, headers, current_sorted)

    # Create Sheet 3: Past Results
    ws_past = wb.create_sheet('Past Results')
    _write_sheet(ws_past, headers, past_sorted)

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return str(output_path)


def _write_sheet(worksheet, headers: List[str], data: List[Dict]):
    """
    Write data to a worksheet with formatting

    Args:
        worksheet: openpyxl worksheet object
        headers: List of column headers
        data: List of dictionaries with resource data
    """
    # Write headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Map headers to data keys
    key_mapping = {
        'Resource ID': None,  # Will be extracted from URL
        'Title': 'title',
        'URL': 'url',
        'Version': 'version',
        'Updated Date': 'updated_date',
        'Developer ID': 'developer_id',
        'Contributor': 'contributor',
        'Tagline': 'tagline'
    }

    # Write data rows
    for row_idx, resource in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            key = key_mapping.get(header)

            # Extract Resource ID from URL
            if header == 'Resource ID':
                value = extract_resource_id(resource.get('url', ''))
            else:
                value = resource.get(key, '')

            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Make URL column clickable
            if header == 'URL' and value:
                cell.hyperlink = value
                cell.font = Font(color='0563C1', underline='single')  # Blue and underlined like a link

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)

        # Set specific widths based on content
        if header == 'Resource ID':
            worksheet.column_dimensions[col_letter].width = 12
        elif header == 'Title':
            worksheet.column_dimensions[col_letter].width = 40
        elif header == 'URL':
            worksheet.column_dimensions[col_letter].width = 50
        elif header == 'Tagline':
            worksheet.column_dimensions[col_letter].width = 60
        elif header == 'Version':
            worksheet.column_dimensions[col_letter].width = 12
        elif header == 'Updated Date':
            worksheet.column_dimensions[col_letter].width = 20
        elif header == 'Developer ID':
            worksheet.column_dimensions[col_letter].width = 15
        elif header == 'Contributor':
            worksheet.column_dimensions[col_letter].width = 25
        else:
            worksheet.column_dimensions[col_letter].width = 20

    # Freeze first row
    worksheet.freeze_panes = 'A2'


def generate_filename(date: datetime = None) -> str:
    """
    Generate filename for Excel report

    Args:
        date: Date to use in filename (defaults to today)

    Returns:
        Filename in format: Ignition Exchange Resource Results_YYMMDD.xlsx
    """
    if date is None:
        date = datetime.now()

    date_str = date.strftime('%y%m%d')
    return f"Ignition Exchange Resource Results_{date_str}.xlsx"
