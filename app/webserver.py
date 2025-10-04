"""
Flask web server for Exchange Scraper dashboard
"""
from flask import Flask, jsonify, request, send_from_directory, send_file
from flask_cors import CORS
from pathlib import Path
import json
import threading
from datetime import datetime
from zoneinfo import ZoneInfo
from typing import Dict, List
import os

# Adelaide timezone
ADELAIDE_TZ = ZoneInfo("Australia/Adelaide")

app = Flask(__name__, static_folder='static')
CORS(app)

# Paths
DATA_DIR = Path('/data')
STATE_FILE = DATA_DIR / 'state.json'
LOG_FILE = DATA_DIR / 'activity.log'
CONFIG_FILE = DATA_DIR / 'config.json'
OUTPUT_DIR = DATA_DIR / 'output'

# Ensure directories exist
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Global state lock
state_lock = threading.Lock()


def get_state() -> Dict:
    """Load current state from file"""
    if STATE_FILE.exists():
        try:
            with open(STATE_FILE, 'r') as f:
                return json.load(f)
        except:
            pass

    # Default state
    return {
        'status': 'idle',
        'progress': {
            'current': 0,
            'total': 0,
            'percentage': 0,
            'current_item': ''
        },
        'current_job': None,
        'last_run': None,
        'next_run': None,
        'elapsed_seconds': 0,
        'estimated_remaining_seconds': 0
    }


def save_state(state: Dict):
    """Save state to file"""
    with state_lock:
        with open(STATE_FILE, 'w') as f:
            json.dump(state, f, indent=2)


def get_config() -> Dict:
    """Load configuration from file"""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r') as f:
                config = json.load(f)
                # Migrate old interval_hours to interval_days
                if 'interval_hours' in config and 'interval_days' not in config:
                    config['interval_days'] = config['interval_hours'] / 24
                    config.pop('interval_hours', None)
                return config
        except:
            pass

    # Default config
    return {
        'interval_days': 7,
        'enabled': True
    }


def save_config(config: Dict):
    """Save configuration to file"""
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=2)


def get_logs(limit: int = 100) -> List[Dict]:
    """Get recent log entries"""
    if not LOG_FILE.exists():
        return []

    try:
        with open(LOG_FILE, 'r') as f:
            lines = f.readlines()

        # Return last N lines
        recent_lines = lines[-limit:] if len(lines) > limit else lines

        logs = []
        for line in recent_lines:
            try:
                log_entry = json.loads(line.strip())
                logs.append(log_entry)
            except:
                pass

        return logs
    except:
        return []


def append_log(message: str, level: str = 'info'):
    """Append log entry to file"""
    log_entry = {
        'timestamp': datetime.now(ADELAIDE_TZ).isoformat(),
        'message': message,
        'level': level
    }

    with open(LOG_FILE, 'a') as f:
        f.write(json.dumps(log_entry) + '\n')


def get_job_history() -> List[Dict]:
    """Get job history"""
    history_file = DATA_DIR / 'job_history.json'

    if history_file.exists():
        try:
            with open(history_file, 'r') as f:
                return json.load(f)
        except:
            pass

    return []


@app.route('/')
def index():
    """Serve dashboard"""
    return send_from_directory('static', 'index.html')


@app.route('/api/status')
def api_status():
    """Get current scraper status"""
    state = get_state()
    return jsonify(state)


@app.route('/api/config', methods=['GET', 'POST'])
def api_config():
    """Get or update configuration"""
    if request.method == 'GET':
        config = get_config()
        return jsonify(config)
    else:
        # Update config
        data = request.json
        config = get_config()

        if 'interval_days' in data:
            config['interval_days'] = int(data['interval_days'])
        # Support legacy interval_hours for backward compatibility
        elif 'interval_hours' in data:
            config['interval_days'] = int(data['interval_hours'] / 24)

        if 'enabled' in data:
            config['enabled'] = bool(data['enabled'])

        # Add notification settings
        if 'notifications' in data:
            config['notifications'] = data['notifications']

        save_config(config)
        append_log(f"Configuration updated: interval={config.get('interval_days', 7)} days")

        return jsonify({'success': True, 'config': config})


@app.route('/api/control/<action>', methods=['POST'])
def api_control(action):
    """Control scraper (run, pause, stop, resume)"""
    state = get_state()

    # Create control signal file for scheduler to read
    control_file = DATA_DIR / f'control_{action}.signal'
    control_file.touch()

    if action == 'run':
        append_log('Manual run triggered', 'info')
        return jsonify({'success': True, 'message': 'Run started'})

    elif action == 'pause':
        append_log('Scraper paused', 'warning')
        return jsonify({'success': True, 'message': 'Paused'})

    elif action == 'stop':
        append_log('Scraper stopped', 'warning')
        return jsonify({'success': True, 'message': 'Stopped'})

    elif action == 'resume':
        append_log('Scraper resumed', 'info')
        return jsonify({'success': True, 'message': 'Resumed'})

    else:
        return jsonify({'success': False, 'message': 'Unknown action'}), 400


@app.route('/api/logs')
def api_logs():
    """Get activity logs"""
    limit = request.args.get('limit', 100, type=int)
    logs = get_logs(limit)
    return jsonify(logs)


@app.route('/api/logs/clear', methods=['POST'])
def api_clear_logs():
    """Clear activity logs"""
    if LOG_FILE.exists():
        LOG_FILE.unlink()

    append_log('Logs cleared', 'info')
    return jsonify({'success': True})


@app.route('/api/history')
def api_history():
    """Get job history"""
    history = get_job_history()
    return jsonify(history)


@app.route('/api/files')
def api_files():
    """List available Excel files"""
    if not OUTPUT_DIR.exists():
        return jsonify([])

    files = []
    for file_path in sorted(OUTPUT_DIR.glob('*.xlsx'), reverse=True):
        stat = file_path.stat()
        files.append({
            'name': file_path.name,
            'size': stat.st_size,
            'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
            'url': f'/api/download/{file_path.name}'
        })

    return jsonify(files)


@app.route('/api/download/<filename>')
def api_download(filename):
    """Download an Excel file"""
    # Security: only allow xlsx files and no path traversal
    if not filename.endswith('.xlsx') or '/' in filename or '\\' in filename:
        return jsonify({'error': 'Invalid filename'}), 400

    file_path = OUTPUT_DIR / filename

    if not file_path.exists():
        return jsonify({'error': 'File not found'}), 404

    return send_file(file_path, as_attachment=True, download_name=filename)


@app.route('/api/delete/<filename>', methods=['DELETE'])
def api_delete_file(filename):
    """Delete an Excel file"""
    # Security: only allow xlsx files and no path traversal
    if not filename.endswith('.xlsx') or '/' in filename or '\\' in filename:
        return jsonify({'error': 'Invalid filename'}), 400

    file_path = OUTPUT_DIR / filename

    if not file_path.exists():
        return jsonify({'error': 'File not found'}), 404

    try:
        file_path.unlink()
        append_log(f'Deleted file: {filename}', 'info')
        return jsonify({'success': True, 'message': f'File {filename} deleted successfully'})
    except Exception as e:
        append_log(f'Failed to delete file {filename}: {str(e)}', 'error')
        return jsonify({'error': f'Failed to delete file: {str(e)}'}), 500


@app.route('/api/test-notification/<channel>', methods=['POST'])
def api_test_notification(channel):
    """Send a test notification to the specified channel (discord, teams, or ntfy)"""
    try:
        import openpyxl
        import sys
        from pathlib import Path as PathLib

        # Add parent directory to path for app imports
        parent_dir = PathLib(__file__).parent.parent
        if str(parent_dir) not in sys.path:
            sys.path.insert(0, str(parent_dir))

        from app.notifications import notify_scrape_complete

        # Find the most recent Excel file
        if not OUTPUT_DIR.exists():
            return jsonify({'error': 'No reports available'}), 404

        excel_files = sorted(OUTPUT_DIR.glob('*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)

        if not excel_files:
            return jsonify({'error': 'No reports available'}), 404

        latest_file = excel_files[0]

        # Read the "Updated Results" sheet
        wb = openpyxl.load_workbook(latest_file, read_only=True)

        if 'Updated Results' not in wb.sheetnames:
            return jsonify({'error': 'No changes sheet found in report'}), 404

        ws_updated = wb['Updated Results']
        ws_current = wb['Current Results'] if 'Current Results' in wb.sheetnames else None
        ws_past = wb['Past Results'] if 'Past Results' in wb.sheetnames else None

        # Read headers and data
        headers = [cell.value for cell in ws_updated[1]]

        # Read updated resources
        updated_resources = []
        for row in ws_updated.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            resource = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    resource[header] = row[i]
            # Convert header names to lowercase with underscores for consistency
            if 'Title' in resource:
                resource['title'] = resource['Title']
            if 'URL' in resource:
                resource['url'] = resource['URL']
            if 'Version' in resource:
                resource['version'] = resource['Version']
            if 'Contributor' in resource:
                resource['contributor'] = resource['Contributor']
            if 'Updated Date' in resource:
                resource['updated_date'] = resource['Updated Date']
            updated_resources.append(resource)

        # Count current and past
        current_count = ws_current.max_row - 1 if ws_current else 0
        past_count = ws_past.max_row - 1 if ws_past else 0

        wb.close()

        # Create stats
        stats = {
            'total_current': current_count,
            'total_past': past_count,
            'total_updated': len(updated_resources),
            'new_count': 0,  # Could be calculated from data
            'modified_count': len(updated_resources)
        }

        # Get config and check if channel is enabled
        config = get_config()
        notifications_config = config.get('notifications', {})

        if channel not in ['discord', 'teams', 'ntfy']:
            return jsonify({'error': 'Invalid channel'}), 400

        if not notifications_config.get(channel, {}).get('enabled'):
            return jsonify({'error': f'{channel.capitalize()} notifications are not enabled'}), 400

        # Send notification to only the specified channel
        append_log(f'Sending test {channel} notification...')

        # Create a config with only the target channel enabled
        test_config = {channel: notifications_config[channel]}

        result = notify_scrape_complete(
            test_config,
            stats,
            latest_file,
            updated_resources
        )

        if result.get(channel):
            append_log(f'Test {channel} notification sent successfully')
            return jsonify({'success': True, 'message': f'Test {channel} notification sent successfully'})
        else:
            append_log(f'Test {channel} notification failed', 'error')
            return jsonify({'error': f'Failed to send test {channel} notification'}), 500

    except Exception as e:
        append_log(f'Test {channel} notification error: {str(e)}', 'error')
        return jsonify({'error': str(e)}), 500


@app.route('/api/changes')
def api_changes():
    """Get latest changes from the most recent Excel report"""
    try:
        import openpyxl
        import sys
        from pathlib import Path

        # Ensure the app directory is in the path for imports
        app_dir = Path(__file__).parent
        if str(app_dir) not in sys.path:
            sys.path.insert(0, str(app_dir))

        # Find the most recent Excel file
        if not OUTPUT_DIR.exists():
            return jsonify({'error': 'No reports available'}), 404

        excel_files = sorted(OUTPUT_DIR.glob('*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)

        if not excel_files:
            return jsonify({'error': 'No reports available'}), 404

        latest_file = excel_files[0]

        # Read the "Updated Results" sheet from the Excel file
        wb = openpyxl.load_workbook(latest_file, read_only=True)

        if 'Updated Results' not in wb.sheetnames:
            return jsonify({'error': 'No changes sheet found in report'}), 404

        ws = wb['Updated Results']

        # Read headers from first row
        headers = [cell.value for cell in ws[1]]

        # Read all data rows
        changes = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Skip empty rows
                continue

            change = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    change[header] = row[i]

            changes.append(change)

        wb.close()

        return jsonify({
            'changes': changes,
            'count': len(changes),
            'file': latest_file.name
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/health')
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy'})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 9089))
    print(f'Starting HTTP web server on port {port}')
    app.run(host='0.0.0.0', port=port, debug=False)
